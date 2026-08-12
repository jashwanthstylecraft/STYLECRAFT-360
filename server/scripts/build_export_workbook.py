#!/usr/bin/env python3
"""Phase 8 export: builds the two-sheet (Graphs, Data) workbook from a JSON
spec written by server/services/exportService.js. This is the one place
that writes the actual .xlsx bytes -- native chart objects bound to Data-
sheet ranges aren't realistically writable from the `xlsx` (SheetJS) package
already used elsewhere in this repo, so this is a separate, explicit runtime
dependency: Python 3 + openpyxl (`pip install openpyxl`). See README's
deployment note for wherever the server ends up hosted.

Usage: python build_export_workbook.py <spec.json> <output.xlsx>
"""
import json
import sys

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.styles import Font

DATA_SHEET = "Data"
GRAPHS_SHEET = "Graphs"
META_MARKER = "#meta"

FIXED_COLUMNS = [
    "metric_slug", "metric_name", "department", "series",
    "row_type", "goal_label", "goal_direction", "aggregation",
]
FIRST_WEEK_COL = len(FIXED_COLUMNS) + 1  # 1-based Excel column index

NUMBER_FORMAT = {
    "currency": "#,##0",
    "percent": "0.0%",
    "decimal": "0.00",
    "count": "0",
}

DEPARTMENT_LABELS = {
    "sales": "Sales",
    "inventory": "Inventory & Purchasing",
    "finance": "Finance",
    "operations": "Operations",
}

PERCENT_LINE_SLUGS = {"weekly-gross-margin", "repair-rate", "in-stock-percentage"}


def write_data_sheet(wb, spec):
    ws = wb.create_sheet(DATA_SHEET)
    weeks = spec["weeks"]
    headers = FIXED_COLUMNS + weeks
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # slug -> {(series, row_type): excel_row_number}
    row_index = {}

    for row in spec["rows"]:
        excel_row = ws.max_row + 1
        fixed = [row[key] for key in ("metric_slug", "metric_name", "department", "series", "row_type", "goal_label", "goal_direction", "aggregation")]
        ws.append(fixed + row["values"])

        number_format = NUMBER_FORMAT.get(row.get("format"), "General")
        for col_offset in range(len(weeks)):
            cell = ws.cell(row=excel_row, column=FIRST_WEEK_COL + col_offset)
            cell.number_format = number_format

        row_index.setdefault(row["metric_slug"], {})[(row["series"], row["row_type"])] = excel_row

    # #meta block — one blank separator row, then key/value rows. Never a
    # merged cell anywhere on this sheet (merges break parsers).
    ws.append([])
    meta = spec["meta"]
    ws.append([META_MARKER])
    for key, value in meta.items():
        ws.append([key, value])

    ws.freeze_panes = ws.cell(row=2, column=FIRST_WEEK_COL).coordinate
    last_week_col = FIRST_WEEK_COL + len(weeks) - 1
    return ws, row_index, last_week_col


def resolve_chart_kind(metric):
    """('line'|'bar', is_stacked) — the Phase 8 spec explicitly wants line
    charts for the three percent-format metrics (a % trend reads better as
    a line than columns), independent of what chart type the live app's
    cards use for them — that's this export's own, deliberately different
    call, not a "fix" to the app's charts."""
    if metric["slug"] in PERCENT_LINE_SLUGS:
        return "line", False
    chart_type = metric.get("chartType")
    if chart_type in ("stacked", "paidUnpaidStacked"):
        return "bar", True
    return "bar", False


def add_series_from_row(chart, ws_data, row_index, slug, series_key, row_type, title, last_week_col):
    excel_row = row_index.get(slug, {}).get((series_key, row_type))
    if excel_row is None:
        return False
    ref = Reference(ws_data, min_col=FIRST_WEEK_COL, max_col=last_week_col, min_row=excel_row, max_row=excel_row)
    series = Series(ref, title=title)
    chart.series.append(series)
    return True


def build_chart(ws_data, row_index, metric, last_week_col, weeks_count):
    kind, stacked = resolve_chart_kind(metric)
    ChartClass = LineChart if kind == "line" else BarChart
    chart = ChartClass()
    chart.title = metric["name"]
    chart.height = 8
    chart.width = 16
    chart.y_axis.numFmt = NUMBER_FORMAT.get(metric.get("format"), "General")
    try:
        chart.displayBlanksAs = "gap"  # never bridge a real gap into a fake trend line
    except AttributeError:
        pass

    if isinstance(chart, BarChart):
        chart.type = "col"
        if stacked:
            chart.grouping = "stacked"
            chart.overlap = 100

    series_keys = metric.get("seriesKeys")
    series_labels = metric.get("seriesLabels")
    any_series = False

    if series_keys:
        for i, key in enumerate(series_keys):
            label = series_labels[i] if series_labels else key
            if add_series_from_row(chart, ws_data, row_index, metric["slug"], key, "value", label, last_week_col):
                any_series = True
        # A metric-level goal duplicated across series rows — only plot it
        # once (from the first series' goal row) so the line isn't drawn
        # twice on top of itself.
        if kind == "bar":
            goal_line = LineChart()
            if add_series_from_row(goal_line, ws_data, row_index, metric["slug"], series_keys[0], "goal", metric.get("goalLabel") or "Goal", last_week_col):
                chart += goal_line
    else:
        if add_series_from_row(chart, ws_data, row_index, metric["slug"], "", "value", metric["name"], last_week_col):
            any_series = True
        if kind == "bar":
            goal_line = LineChart()
            if add_series_from_row(goal_line, ws_data, row_index, metric["slug"], "", "goal", metric.get("goalLabel") or "Goal", last_week_col):
                chart += goal_line
        elif kind == "line":
            if add_series_from_row(chart, ws_data, row_index, metric["slug"], "", "goal", metric.get("goalLabel") or "Goal", last_week_col):
                pass

    if not any_series:
        return None

    cats_ref = Reference(ws_data, min_col=FIRST_WEEK_COL, max_col=last_week_col, min_row=1, max_row=1)
    chart.set_categories(cats_ref)
    chart.legend = None if not series_keys else chart.legend
    return chart


def write_graphs_sheet(wb, spec, ws_data, row_index, last_week_col):
    ws = wb.create_sheet(GRAPHS_SHEET)
    ws["A1"] = "StyleCraft 360 — Data Export"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "The whole business, at a glance."
    meta = spec["meta"]
    ws["A3"] = f"Exported {meta['exportedAt']} · Weeks {meta['weekRangeFrom']} to {meta['weekRangeTo']}"

    weeks_count = len(spec["weeks"])
    ROWS_PER_CHART = 17
    COLS_PER_CHART = 9
    cursor_row = 6

    by_department = {}
    for metric in spec["chartMetrics"]:
        by_department.setdefault(metric["department"], []).append(metric)

    for department_key in ("sales", "inventory", "finance", "operations"):
        metrics = by_department.get(department_key, [])
        if not metrics:
            continue
        header_cell = ws.cell(row=cursor_row, column=1, value=DEPARTMENT_LABELS.get(department_key, department_key))
        header_cell.font = Font(bold=True, size=12)
        cursor_row += 2

        col_slot = 0
        row_at_slot_start = cursor_row
        for metric in metrics:
            chart = build_chart(ws_data, row_index, metric, last_week_col, weeks_count)
            if chart is not None:
                anchor_col_letter = ws.cell(row=1, column=1 + col_slot * COLS_PER_CHART).coordinate[0:-1] or chr(ord("A") + col_slot * COLS_PER_CHART)
                anchor = f"{ws.cell(row=row_at_slot_start, column=1 + col_slot * COLS_PER_CHART).coordinate}"
                ws.add_chart(chart, anchor)
            col_slot += 1
            if col_slot >= 2:
                col_slot = 0
                row_at_slot_start += ROWS_PER_CHART

        cursor_row = row_at_slot_start + (ROWS_PER_CHART if col_slot != 0 else 0) + 2

    return ws


def main():
    spec_path, output_path = sys.argv[1], sys.argv[2]
    with open(spec_path, "r", encoding="utf-8") as f:
        spec = json.load(f)

    wb = Workbook()
    wb.remove(wb.active)  # default blank sheet

    # Graphs is written FIRST so it's the leftmost/active tab, but it needs
    # the Data sheet's cells to already exist to reference them — write Data
    # to the workbook first, then reorder tabs afterward.
    ws_data, row_index, last_week_col = write_data_sheet(wb, spec)
    write_graphs_sheet(wb, spec, ws_data, row_index, last_week_col)
    wb.move_sheet(GRAPHS_SHEET, offset=-1)  # Graphs before Data in tab order
    wb.active = 0

    wb.save(output_path)


if __name__ == "__main__":
    main()
